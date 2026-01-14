# app.py
# iServer Reporting DB — Hierarchy SQL Generator (Streamlit)
# Fix: RDL schema (2016) requires ReportSections -> ReportSection -> Body/Page/Width (Body cannot be direct child of Report)
# Features:
# - Multiple relationship types per hop
# - Bi-Directional relationship option (default)
# - Execute SQL safely in-app (ORDER BY applied outside wrapper)
# - Excel export with hierarchy merges (row-span) + merged preview (HTML rowspan)
# - RDL (Report Builder) export (production-ready):
#     * RTL/LTR option (RTL default)
#     * Better header styles + repeating headers
#     * Hierarchy display via Row Groups + suppression of repeated group values
#     * Auto column widths from last executed dataset
#     * Auto fit to page width

import re
from io import BytesIO
from dataclasses import dataclass, field
from typing import List, Optional, Dict, Any, Tuple

import streamlit as st
import pyodbc
import pandas as pd

from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

import xml.etree.ElementTree as ET
from xml.dom import minidom


st.set_page_config(page_title="iServer Hierarchy SQL Generator", layout="wide")


# ---------------------------
# Utilities
# ---------------------------
def qname(schema: str, name: str) -> str:
    """Return [schema].[name] with safe brackets."""
    def wrap(x: str) -> str:
        x = x.strip()
        if x.startswith("[") and x.endswith("]"):
            return x
        return f"[{x.replace(']', ']]')}]"
    return f"{wrap(schema)}.{wrap(name)}"


def escape_sql_string(v: str) -> str:
    return v.replace("'", "''")


def build_condition(column_sql: str, op: str, raw_values: str) -> str:
    """
    raw_values rules:
      - CONTAINS/STARTS/ENDS/LIKE: raw_values is single string
      - = / <> / >= / <= / > / <: raw_values single string
      - IN / NOT IN: comma-separated values (strings)
    """
    op = op.upper().strip()
    raw_values = (raw_values or "").strip()

    if op in ("LIKE", "CONTAINS", "STARTS WITH", "ENDS WITH"):
        val = escape_sql_string(raw_values)
        if op == "LIKE":
            pattern = val
        elif op == "CONTAINS":
            pattern = f"%{val}%"
        elif op == "STARTS WITH":
            pattern = f"{val}%"
        else:  # ENDS WITH
            pattern = f"%{val}"
        return f"{column_sql} LIKE N'{pattern}'"

    if op in ("IN", "NOT IN"):
        parts = [p.strip() for p in raw_values.split(",") if p.strip()]
        if not parts:
            return "1=1"
        parts_escaped = [f"N'{escape_sql_string(p)}'" for p in parts]
        return f"{column_sql} {op} ({', '.join(parts_escaped)})"

    if not raw_values:
        return "1=1"
    val = escape_sql_string(raw_values)
    return f"{column_sql} {op} N'{val}'"


def fetch_df(conn: pyodbc.Connection, sql: str, params: Optional[List[Any]] = None) -> pd.DataFrame:
    return pd.read_sql(sql, conn, params=params or [])


def list_views(conn: pyodbc.Connection, schema: str) -> List[str]:
    sql = """
    SELECT v.name AS view_name
    FROM sys.views v
    JOIN sys.schemas s ON s.schema_id = v.schema_id
    WHERE s.name = ?
    ORDER BY v.name
    """
    df = fetch_df(conn, sql, [schema])
    return df["view_name"].tolist()


def list_columns(conn: pyodbc.Connection, schema: str, view_name: str) -> List[str]:
    full = f"{schema}.{view_name}"
    sql = """
    SELECT c.name AS column_name
    FROM sys.columns c
    WHERE c.object_id = OBJECT_ID(?)
    ORDER BY c.column_id
    """
    df = fetch_df(conn, sql, [full])
    return df["column_name"].tolist()


def make_unique_alias(base_alias: str, level_tag: str, used: set) -> str:
    """Ensure output alias is unique (required for derived-table wrapper)."""
    base_alias = (base_alias or "").strip() or "عمود"
    level_tag = (level_tag or "").strip()

    candidate = base_alias
    if candidate in used:
        candidate = f"{base_alias} - {level_tag}" if level_tag else f"{base_alias} - مستوى"

    n = 2
    while candidate in used:
        candidate = f"{base_alias} - {level_tag} ({n})" if level_tag else f"{base_alias} ({n})"
        n += 1

    used.add(candidate)
    return candidate


def parse_conn_kv(conn_str: str) -> Dict[str, str]:
    """Parse ODBC connection string into dict (best-effort)."""
    kv = {}
    for part in (conn_str or "").split(";"):
        part = part.strip()
        if not part or "=" not in part:
            continue
        k, v = part.split("=", 1)
        kv[k.strip().lower()] = v.strip()
    return kv


# ---------------------------
# Excel merge helpers (row-span)
# ---------------------------
def apply_hierarchy_merges(ws, df: pd.DataFrame, merge_cols: List[str]):
    """
    Merge repeated values vertically in Excel, respecting hierarchy:
    a merge in column i only happens when all previous merge columns (0..i-1) are equal too.
    Assumes df is written with headers in row 1 and data starts row 2.
    """
    if df.empty or not merge_cols:
        return

    merge_cols = [c for c in merge_cols if c in df.columns]
    if not merge_cols:
        return

    n = len(df)

    for col_pos, col_name in enumerate(merge_cols):
        r = 0
        while r < n:
            start = r
            val = df.iloc[start][col_name]

            r2 = r + 1
            while r2 < n:
                if df.iloc[r2][col_name] != val:
                    break

                same_prefix = True
                for k in range(col_pos):
                    if df.iloc[r2][merge_cols[k]] != df.iloc[start][merge_cols[k]]:
                        same_prefix = False
                        break

                if not same_prefix:
                    break
                r2 += 1

            end = r2 - 1

            if end > start and pd.notna(val):
                excel_row_start = start + 2
                excel_row_end = end + 2
                excel_col = df.columns.get_loc(col_name) + 1

                ws.merge_cells(
                    start_row=excel_row_start, start_column=excel_col,
                    end_row=excel_row_end, end_column=excel_col
                )
                cell = ws.cell(excel_row_start, excel_col)
                cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

            r = r2


def autosize_columns(ws, df: pd.DataFrame, max_width: int = 60):
    """Auto-size Excel columns based on content length (simple heuristic)."""
    for j, col in enumerate(df.columns, start=1):
        col_letter = get_column_letter(j)
        max_len = len(str(col)) if col is not None else 0
        for i in range(min(len(df), 2000)):  # cap scan for performance
            v = df.iloc[i, j - 1]
            if pd.notna(v):
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), max_width)


def df_to_excel_bytes(
    df: pd.DataFrame,
    sheet_name: str = "Result",
    merge_hierarchy: bool = False,
    merge_cols: Optional[List[str]] = None
) -> bytes:
    """Export df to Excel. Optionally merges repeated hierarchy cells vertically."""
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        # Header styling
        header_font = Font(bold=True)
        for j in range(1, len(df.columns) + 1):
            c = ws.cell(row=1, column=j)
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Normal cell alignment
        for row in ws.iter_rows(min_row=2, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        if merge_hierarchy and merge_cols:
            apply_hierarchy_merges(ws, df, merge_cols)

        autosize_columns(ws, df)

    return bio.getvalue()


# ---------------------------
# Rowspan (HTML preview) helpers
# ---------------------------
def compute_rowspans(df: pd.DataFrame, merge_cols: List[str]):
    """
    Compute rowspan info for HTML table merging with hierarchy rules.
    Returns dict: (row_idx, col_name) -> rowspan, and a set of skipped cells.
    """
    rowspans = {}
    skip = set()
    if df.empty or not merge_cols:
        return rowspans, skip

    merge_cols = [c for c in merge_cols if c in df.columns]
    n = len(df)

    for col_pos, col_name in enumerate(merge_cols):
        r = 0
        while r < n:
            start = r
            val = df.iloc[start][col_name]
            r2 = r + 1
            while r2 < n:
                if df.iloc[r2][col_name] != val:
                    break
                same_prefix = True
                for k in range(col_pos):
                    if df.iloc[r2][merge_cols[k]] != df.iloc[start][merge_cols[k]]:
                        same_prefix = False
                        break
                if not same_prefix:
                    break
                r2 += 1
            end = r2 - 1
            span = end - start + 1
            if span > 1 and pd.notna(val):
                rowspans[(start, col_name)] = span
                for rr in range(start + 1, end + 1):
                    skip.add((rr, col_name))
            r = r2

    return rowspans, skip


def df_to_rowspan_html(df: pd.DataFrame, merge_cols: List[str], rtl: bool = True) -> str:
    """Render df as HTML table with rowspan merges (hierarchy-aware)."""
    rowspans, skip = compute_rowspans(df, merge_cols)

    direction = "rtl" if rtl else "ltr"
    align = "right" if rtl else "left"

    style = f"""
    <style>
      table.hier {{ border-collapse: collapse; width: 100%; direction: {direction}; }}
      table.hier th, table.hier td {{ border: 1px solid #ddd; padding: 6px; vertical-align: top; }}
      table.hier th {{ font-weight: 700; text-align: center; background: #f2f4f8; }}
      table.hier td {{ text-align: {align}; white-space: pre-wrap; }}
    </style>
    """

    html = [style, '<table class="hier">', "<thead><tr>"]
    for c in df.columns:
        html.append(f"<th>{c}</th>")
    html.append("</tr></thead><tbody>")

    for i in range(len(df)):
        html.append("<tr>")
        for c in df.columns:
            if (i, c) in skip:
                continue
            v = df.iloc[i][c]
            v_str = "" if pd.isna(v) else str(v)
            span = rowspans.get((i, c), 1)
            if span > 1:
                html.append(f'<td rowspan="{span}">{v_str}</td>')
            else:
                html.append(f"<td>{v_str}</td>")
        html.append("</tr>")

    html.append("</tbody></table>")
    return "".join(html)


# ---------------------------
# Data Models
# ---------------------------
@dataclass
class LevelCondition:
    column: str
    operator: str
    values: str


@dataclass
class LevelSpec:
    level_no: int
    object_view: str  # ShapeType view name
    level_alias_ar: str  # label used to help uniqueness
    columns: List[str] = field(default_factory=list)
    col_alias_ar: Dict[str, str] = field(default_factory=dict)
    conditions: List[LevelCondition] = field(default_factory=list)


@dataclass
class HopSpec:
    from_level_no: int
    relationship_views: List[str]  # multiple RelationshipType views
    join_direction: str            # "Bi-Directional" / "From->To" / "To->From"


# ---------------------------
# SQL Generator
# ---------------------------
def generate_sql_parts(
    levels: List[LevelSpec],
    hops: List[HopSpec],
    shape_schema: str = "ShapeType",
    rel_schema: str = "RelationshipType",
    distinct: bool = True,
    include_order_by: bool = True
) -> Tuple[str, str, List[str], List[str]]:
    """
    Returns:
      sql_full: standalone SQL (ORDER BY uses Lx expressions) for SSMS
      sql_core: same SQL without ORDER BY (safe to wrap)
      wrapper_order_cols: list of X.[<unique alias>] to use in outer ORDER BY
      output_fields: list of output column aliases as produced in SELECT (for RDL)
    """
    if not levels:
        return "-- No levels defined.", "-- No levels defined.", [], []

    lvl_alias = {lvl.level_no: f"L{lvl.level_no}" for lvl in levels}
    levels_by_no = {lvl.level_no: lvl for lvl in levels}

    used_output_aliases = set()
    select_cols = []
    output_fields = []
    first_output_alias_per_level: Dict[int, str] = {}

    for lvl in levels:
        a = lvl_alias[lvl.level_no]
        level_tag = (lvl.level_alias_ar or f"Level {lvl.level_no}").strip()

        for col in lvl.columns:
            base_ar = (lvl.col_alias_ar.get(col) or col).strip()
            unique_ar = make_unique_alias(base_ar, level_tag, used_output_aliases)

            if lvl.level_no not in first_output_alias_per_level:
                first_output_alias_per_level[lvl.level_no] = unique_ar

            select_cols.append(f"{a}.[{col}] AS [{unique_ar}]")
            output_fields.append(unique_ar)

    if not select_cols:
        l1 = levels[0]
        a = lvl_alias[l1.level_no]
        unique_ar = make_unique_alias("ID", (l1.level_alias_ar or "Level 1"), used_output_aliases)
        first_output_alias_per_level[l1.level_no] = unique_ar
        select_cols = [f"{a}.[ShapeID] AS [{unique_ar}]"]
        output_fields = [unique_ar]

    distinct_sql = "DISTINCT " if distinct else ""

    l1 = levels[0]
    from_sql = f"FROM {qname(shape_schema, l1.object_view)} {lvl_alias[l1.level_no]}"

    join_sql_parts = []
    where_parts = []

    for cond in l1.conditions:
        where_parts.append(build_condition(f"{lvl_alias[l1.level_no]}.[{cond.column}]", cond.operator, cond.values))

    for hop in hops:
        i = hop.from_level_no
        j = i + 1
        Ai = lvl_alias[i]
        Aj = lvl_alias[j]
        R = f"R{i}_{j}"

        if j not in levels_by_no:
            continue

        next_obj_full = qname(shape_schema, levels_by_no[j].object_view)

        if not hop.relationship_views:
            return (
                f"-- Please select at least one Relationship between Level {i} and Level {j}.",
                f"-- Please select at least one Relationship between Level {i} and Level {j}.",
                [],
                output_fields
            )

        # Derived relationship set (supports multi-view + bi-directional)
        rel_union_parts = []
        for rv in hop.relationship_views:
            rel_union_parts.append(f"SELECT FromObjectID, ToObjectID FROM {qname(rel_schema, rv)}")
            if hop.join_direction == "Bi-Directional":
                rel_union_parts.append(
                    f"SELECT ToObjectID AS FromObjectID, FromObjectID AS ToObjectID FROM {qname(rel_schema, rv)}"
                )

        rel_derived = (
            "(\n"
            "  SELECT DISTINCT FromObjectID, ToObjectID\n"
            "  FROM (\n    " + "\n    UNION ALL\n    ".join(rel_union_parts) + "\n"
            "  ) U\n"
            ")"
        )

        if hop.join_direction in ("From->To", "Bi-Directional"):
            join_sql_parts.append(f"INNER JOIN {rel_derived} {R} ON {Ai}.[ShapeID] = {R}.[FromObjectID]")
            join_sql_parts.append(f"INNER JOIN {next_obj_full} {Aj} ON {Aj}.[ShapeID] = {R}.[ToObjectID]")
        else:  # "To->From"
            join_sql_parts.append(f"INNER JOIN {rel_derived} {R} ON {Ai}.[ShapeID] = {R}.[ToObjectID]")
            join_sql_parts.append(f"INNER JOIN {next_obj_full} {Aj} ON {Aj}.[ShapeID] = {R}.[FromObjectID]")

        for cond in levels_by_no[j].conditions:
            where_parts.append(build_condition(f"{Aj}.[{cond.column}]", cond.operator, cond.values))

    where_sql = ""
    if where_parts:
        where_sql = "WHERE " + "\n  AND ".join(where_parts)

    # SSMS ORDER BY using Lx expressions
    order_sql_ssms = ""
    if include_order_by:
        order_exprs = []
        for lvl in levels:
            a = lvl_alias[lvl.level_no]
            if lvl.columns:
                order_exprs.append(f"{a}.[{lvl.columns[0]}]")
        if order_exprs:
            order_sql_ssms = "ORDER BY " + ", ".join(order_exprs)

    # Wrapper ORDER BY using output aliases (safe)
    wrapper_order_cols = []
    for lvl in levels:
        out_alias = first_output_alias_per_level.get(lvl.level_no)
        if out_alias:
            wrapper_order_cols.append(f"X.[{out_alias}]")

    sql_core = (
        f"SELECT {distinct_sql}\n  " + ",\n  ".join(select_cols) + "\n"
        + from_sql + "\n"
        + ("\n".join(join_sql_parts) + "\n" if join_sql_parts else "")
        + where_sql
    ).strip()

    sql_full = sql_core
    if order_sql_ssms:
        sql_full = (sql_full + "\n" + order_sql_ssms).strip()

    return sql_full, sql_core, wrapper_order_cols, output_fields


# ---------------------------
# RDL (Report Builder) Generator — FIXED SCHEMA (ReportSections)
# ---------------------------
def _prettify_xml(elem: ET.Element) -> bytes:
    rough = ET.tostring(elem, encoding="utf-8", xml_declaration=True)
    parsed = minidom.parseString(rough)
    return parsed.toprettyxml(indent="  ", encoding="utf-8")


def _calc_col_widths_cm_from_df(
    df: Optional[pd.DataFrame],
    fields: List[str],
    usable_width_cm: float,
    sample_rows: int = 2000,
    min_cm: float = 1.8,
    max_cm: float = 6.2
) -> List[float]:
    """Estimate column widths from content lengths and fit into usable page width."""
    n = max(1, len(fields))
    if df is None or df.empty:
        w = usable_width_cm / n
        w = max(min_cm, min(max_cm, w))
        return [w] * n

    scan = df.head(sample_rows)
    lens = []
    for f in fields:
        max_len = len(str(f))
        if f in scan.columns:
            col = scan[f].astype(str)
            max_len = max(max_len, int(col.map(len).max()))
        max_len = min(max_len, 80)
        lens.append(max_len)

    weights = [max(1.0, float(l)) for l in lens]
    total = sum(weights)
    raw = [usable_width_cm * (w / total) for w in weights]

    clamped = [max(min_cm, min(max_cm, x)) for x in raw]
    sum_clamped = sum(clamped)

    if sum_clamped > usable_width_cm:
        scale = usable_width_cm / sum_clamped
        clamped = [max(min_cm, x * scale) for x in clamped]

    sum_clamped = sum(clamped)
    if sum_clamped < usable_width_cm * 0.98:
        scale = min(1.30, usable_width_cm / max(sum_clamped, 0.001))
        clamped = [min(max_cm, x * scale) for x in clamped]

    return clamped


def generate_rdl_bytes(
    report_name: str,
    sql_query: str,
    fields: List[str],
    server: str,
    database: str,
    group_cols: Optional[List[str]] = None,
    rtl: bool = True,
    page_width_cm: float = 21.0,
    page_height_cm: float = 29.7,
    margin_cm: float = 1.0,
    df_for_sizing: Optional[pd.DataFrame] = None,
    sample_rows_for_sizing: int = 2000,
    auto_fit_page_width: bool = True,
    repeat_headers: bool = True
) -> bytes:
    """
    Valid SSRS 2016 RDL:
      Report -> DataSources, DataSets, ReportSections -> ReportSection -> Body + Width + Page
    Notes:
      - NO TablixCorner emitted (avoids empty corner errors)
      - Tablix element ordering is aligned with Report Builder expectations
    """

    fields = list(fields or [])
    if not fields:
        fields = ["Result"]

    group_cols = [c for c in (group_cols or []) if c in fields]

    ns = "http://schemas.microsoft.com/sqlserver/reporting/2016/01/reportdefinition"
    rds = "http://schemas.microsoft.com/SQLServer/reporting/reportdesigner"
    ET.register_namespace("", ns)
    ET.register_namespace("rd", rds)

    def E(tag: str) -> str:
        return f"{{{ns}}}{tag}"

    def RD(tag: str) -> str:
        return f"{{{rds}}}{tag}"

    def _add_style(parent: ET.Element, style_dict: Dict[str, str]):
        st_el = ET.SubElement(parent, E("Style"))
        for k, v in style_dict.items():
            if k.startswith("Border."):
                border = st_el.find(E("Border"))
                if border is None:
                    border = ET.SubElement(st_el, E("Border"))
                sub = k.split(".", 1)[1]
                ET.SubElement(border, E(sub)).text = v
            else:
                ET.SubElement(st_el, E(k)).text = v
        return st_el

    def textbox(name: str, value_expr: str, is_header: bool = False, rtl_: bool = True):
        tb = ET.Element(E("Textbox"), Name=name)
        ET.SubElement(tb, E("CanGrow")).text = "true"
        ET.SubElement(tb, E("KeepTogether")).text = "true"

        paragraphs = ET.SubElement(tb, E("Paragraphs"))
        par = ET.SubElement(paragraphs, E("Paragraph"))
        textruns = ET.SubElement(par, E("TextRuns"))
        run = ET.SubElement(textruns, E("TextRun"))
        ET.SubElement(run, E("Value")).text = value_expr

        _add_style(run, {
            "FontSize": "9pt" if not is_header else "9.5pt",
            "FontWeight": "Bold" if is_header else "Normal",
            "Color": "#111827",
        })
        ET.SubElement(par, E("Style"))

        align = "Right" if rtl_ else "Left"
        bg = "#EEF2FF" if is_header else "White"
        _add_style(tb, {
            "TextAlign": "Center" if is_header else align,
            "VerticalAlign": "Middle" if is_header else "Top",
            "BackgroundColor": bg,
            "PaddingLeft": "4pt",
            "PaddingRight": "4pt",
            "PaddingTop": "2pt",
            "PaddingBottom": "2pt",
            "Border.Color": "#CBD5E1",
            "Border.Style": "Solid",
            "Border.Width": "0.6pt",
        })
        return tb

    # ---------------- Report root ----------------
    report = ET.Element(E("Report"))
    ET.SubElement(report, E("Language")).text = "ar-QA" if rtl else "en-US"
    ET.SubElement(report, E("ConsumeContainerWhitespace")).text = "true"

    # ---------------- DataSources ----------------
    datasources = ET.SubElement(report, E("DataSources"))
    ds = ET.SubElement(datasources, E("DataSource"), Name="iServerDS")
    connprops = ET.SubElement(ds, E("ConnectionProperties"))
    ET.SubElement(connprops, E("DataProvider")).text = "SQL"
    ET.SubElement(connprops, E("ConnectString")).text = (
        f"Data Source={server};Initial Catalog={database};Integrated Security=True;"
    )
    ET.SubElement(ds, RD("DataSourceID")).text = "00000000-0000-0000-0000-000000000000"

    # ---------------- DataSets ----------------
    datasets = ET.SubElement(report, E("DataSets"))
    dset = ET.SubElement(datasets, E("DataSet"), Name="MainDS")
    query = ET.SubElement(dset, E("Query"))
    ET.SubElement(query, E("DataSourceName")).text = "iServerDS"
    ET.SubElement(query, E("CommandText")).text = sql_query

    fields_el = ET.SubElement(dset, E("Fields"))
    for f in fields:
        fld = ET.SubElement(fields_el, E("Field"), Name=f)
        ET.SubElement(fld, E("DataField")).text = f
        ET.SubElement(fld, RD("TypeName")).text = "System.String"

    # ---------------- ReportSections ----------------
    report_sections = ET.SubElement(report, E("ReportSections"))
    section = ET.SubElement(report_sections, E("ReportSection"))

    # sizing
    usable_width_cm = page_width_cm - (2 * margin_cm)
    if not auto_fit_page_width:
        usable_width_cm = min(usable_width_cm, 18.0)

    col_widths = _calc_col_widths_cm_from_df(
        df_for_sizing, fields, usable_width_cm,
        sample_rows=sample_rows_for_sizing
    )
    tablix_width_cm = sum(col_widths)

    # ---------------- Body (under ReportSection) ----------------
    body = ET.SubElement(section, E("Body"))
    reportitems = ET.SubElement(body, E("ReportItems"))

    # Title
    title = ET.SubElement(reportitems, E("Textbox"), Name="Title")
    ET.SubElement(title, E("CanGrow")).text = "true"
    ET.SubElement(title, E("KeepTogether")).text = "true"
    pars = ET.SubElement(title, E("Paragraphs"))
    par = ET.SubElement(pars, E("Paragraph"))
    trs = ET.SubElement(par, E("TextRuns"))
    tr = ET.SubElement(trs, E("TextRun"))
    ET.SubElement(tr, E("Value")).text = report_name
    _add_style(tr, {"FontSize": "14pt", "FontWeight": "Bold", "Color": "#0F172A"})
    ET.SubElement(par, E("Style"))
    _add_style(title, {"TextAlign": "Right" if rtl else "Left", "PaddingLeft": "4pt", "PaddingRight": "4pt"})
    ET.SubElement(title, E("Top")).text = "0.2cm"
    ET.SubElement(title, E("Left")).text = "0.2cm"
    ET.SubElement(title, E("Height")).text = "0.9cm"
    ET.SubElement(title, E("Width")).text = f"{usable_width_cm:.2f}cm"

    # ---------------- Tablix (ORDER MATTERS) ----------------
    tablix = ET.SubElement(reportitems, E("Tablix"), Name="Tablix1")

    # TablixBody
    tablixbody = ET.SubElement(tablix, E("TablixBody"))

    # Columns
    tablixcols = ET.SubElement(tablixbody, E("TablixColumns"))
    for w in col_widths:
        tc = ET.SubElement(tablixcols, E("TablixColumn"))
        ET.SubElement(tc, E("Width")).text = f"{w:.2f}cm"

    # Rows
    tablixrows = ET.SubElement(tablixbody, E("TablixRows"))

    def add_row(items: List[ET.Element], height_cm: float):
        row = ET.SubElement(tablixrows, E("TablixRow"))
        ET.SubElement(row, E("Height")).text = f"{height_cm:.2f}cm"
        rowcells = ET.SubElement(row, E("TablixCells"))
        for item in items:
            cell = ET.SubElement(rowcells, E("TablixCell"))
            contents = ET.SubElement(cell, E("CellContents"))
            contents.append(item)

    # Header row
    header_cells = [textbox(f"Hdr_{i}", f'"{col}"', is_header=True, rtl_=rtl) for i, col in enumerate(fields)]
    add_row(header_cells, 0.85)

    # Detail row + group suppression (merge-like)
    detail_cells = []
    for i, col in enumerate(fields):
        if col in group_cols:
            expr = f'=IIF(RowNumber("Grp_{col}") = 1, Fields!{col}.Value, Nothing)'
        else:
            expr = f"=Fields!{col}.Value"
        detail_cells.append(textbox(f"Cell_{i}", expr, is_header=False, rtl_=rtl))
    add_row(detail_cells, 0.75)

    # Column hierarchy
    colhier = ET.SubElement(tablix, E("TablixColumnHierarchy"))
    colmembers = ET.SubElement(colhier, E("TablixMembers"))
    for _ in fields:
        ET.SubElement(colmembers, E("TablixMember"))

    # Row hierarchy
    rowhier = ET.SubElement(tablix, E("TablixRowHierarchy"))
    rowmembers = ET.SubElement(rowhier, E("TablixMembers"))

    # Header static member
    header_member = ET.SubElement(rowmembers, E("TablixMember"))
    if repeat_headers:
        ET.SubElement(header_member, E("RepeatOnNewPage")).text = "true"
        ET.SubElement(header_member, E("KeepWithGroup")).text = "After"
        ET.SubElement(header_member, E("FixedData")).text = "true"

    # Groups then detail member
    if group_cols:
        parent = rowmembers
        for gcol in group_cols:
            gm = ET.SubElement(parent, E("TablixMember"))
            grp = ET.SubElement(gm, E("Group"), Name=f"Grp_{gcol}")
            exprs = ET.SubElement(grp, E("GroupExpressions"))
            ET.SubElement(exprs, E("GroupExpression")).text = f"=Fields!{gcol}.Value"
            ET.SubElement(gm, E("KeepTogether")).text = "true"
            child = ET.SubElement(gm, E("TablixMembers"))
            parent = child
        ET.SubElement(parent, E("TablixMember"))  # detail leaf
    else:
        ET.SubElement(rowmembers, E("TablixMember"))

    # Now remaining Tablix props AFTER hierarchies
    ET.SubElement(tablix, E("DataSetName")).text = "MainDS"
    ET.SubElement(tablix, E("Top")).text = "1.4cm"
    ET.SubElement(tablix, E("Left")).text = "0.2cm"
    ET.SubElement(tablix, E("Height")).text = "5cm"
    ET.SubElement(tablix, E("Width")).text = f"{tablix_width_cm:.2f}cm"
    _add_style(tablix, {"Direction": "RTL" if rtl else "LTR"})

    # Body end
    ET.SubElement(body, E("Height")).text = "7cm"
    ET.SubElement(body, E("Style"))

    # Section width + Page
    ET.SubElement(section, E("Width")).text = f"{usable_width_cm:.2f}cm"

    page = ET.SubElement(section, E("Page"))
    ET.SubElement(page, E("PageHeight")).text = f"{page_height_cm:.1f}cm"
    ET.SubElement(page, E("PageWidth")).text = f"{page_width_cm:.1f}cm"
    ET.SubElement(page, E("LeftMargin")).text = f"{margin_cm:.1f}cm"
    ET.SubElement(page, E("RightMargin")).text = f"{margin_cm:.1f}cm"
    ET.SubElement(page, E("TopMargin")).text = f"{margin_cm:.1f}cm"
    ET.SubElement(page, E("BottomMargin")).text = f"{margin_cm:.1f}cm"

    # Designer metadata
    ET.SubElement(report, RD("ReportUnitType")).text = "Cm"
    ET.SubElement(report, RD("ReportID")).text = "11111111-1111-1111-1111-111111111111"

    return _prettify_xml(report)


# ---------------------------
# UI
# ---------------------------
st.title("iServer Reporting DB — Hierarchy SQL Generator")

DEFAULT_TEMPLATE = "Driver={ODBC Driver 17 for SQL Server};Server=[Server];Database=[Database];Trusted_Connection=yes;TrustServerCertificate=yes;"

with st.expander("1) DB Connection", expanded=True):
    st.info("Default uses Windows Authentication (Trusted_Connection).")

    override_full = st.checkbox("Override with full connection string", value=False, key="override_full")

    if override_full:
        conn_str = st.text_input(
            "Full ODBC Connection String",
            value=DEFAULT_TEMPLATE,
            type="password",
            key="conn_full"
        )
    else:
        c1, c2 = st.columns(2)
        with c1:
            server = st.text_input("Server", placeholder="e.g. malvmiservdb22", key="server")
        with c2:
            database = st.text_input("Database", placeholder="e.g. iServerReportingDB_DEV", key="database")

        st.caption(f"Template: {DEFAULT_TEMPLATE}")
        conn_str = DEFAULT_TEMPLATE.replace("[Server]", server.strip()).replace("[Database]", database.strip())

    connect_btn = st.button("Connect", type="primary")

# Session defaults
for k, v in {
    "conn_ok": False,
    "shape_views": [],
    "rel_views": [],
    "level_count": 3,
    "last_result_df": None,
    "last_exec_sql": None,
}.items():
    st.session_state.setdefault(k, v)

conn: Optional[pyodbc.Connection] = None

if connect_btn:
    try:
        if not override_full:
            if not (st.session_state.get("server", "").strip() and st.session_state.get("database", "").strip()):
                st.session_state.conn_ok = False
                st.error("Please fill Server and Database.")
            else:
                conn = pyodbc.connect(conn_str, timeout=8)
                st.session_state.conn_ok = True
        else:
            if not conn_str.strip():
                st.session_state.conn_ok = False
                st.error("Please provide a connection string.")
            else:
                conn = pyodbc.connect(conn_str, timeout=8)
                st.session_state.conn_ok = True

        if st.session_state.conn_ok:
            st.success("Connected ✅")
            st.session_state.shape_views = list_views(conn, "ShapeType")
            st.session_state.rel_views = list_views(conn, "RelationshipType")

    except Exception as e:
        st.session_state.conn_ok = False
        st.error(f"Connection failed: {e}")

if st.session_state.conn_ok:
    try:
        if conn is None:
            conn = pyodbc.connect(conn_str, timeout=8)
    except Exception:
        conn = None

    st.divider()
    st.subheader("2) Define Hierarchy Levels")

    st.session_state.level_count = st.number_input(
        "Number of levels",
        min_value=1, max_value=10,
        value=st.session_state.level_count
    )

    levels: List[LevelSpec] = []
    hops: List[HopSpec] = []

    for i in range(1, int(st.session_state.level_count) + 1):
        with st.expander(f"Level {i}", expanded=(i <= 2)):
            obj_view = st.selectbox(
                f"Object/View for Level {i} (ShapeType.[...])",
                options=st.session_state.shape_views,
                key=f"lvl{i}_obj"
            )

            lvl_alias_ar = st.text_input(
                f"Arabic label for Level {i} (used to auto-unique column aliases)",
                value="",
                key=f"lvl{i}_label"
            )

            cols = list_columns(conn, "ShapeType", obj_view) if conn else []
            default_cols = [c for c in ["ShapeName", "ShapeID"] if c in cols]
            selected_cols = st.multiselect(
                f"Columns to output from Level {i}",
                options=cols,
                default=default_cols,
                key=f"lvl{i}_cols"
            )

            st.markdown("**Aliases for selected columns**")
            col_alias_ar: Dict[str, str] = {}
            for c in selected_cols:
                default_ar = {"ShapeID": "المعرّف", "ShapeName": "الاسم"}.get(c, c)
                col_alias_ar[c] = st.text_input(
                    f"Alias for '{c}'",
                    value=default_ar,
                    key=f"lvl{i}_alias_{c}"
                )

            st.markdown("**Conditions (optional)**")
            st.session_state.setdefault(f"lvl{i}_cond_count", 0)
            ccount = st.number_input(
                f"Number of conditions for Level {i}",
                min_value=0, max_value=10,
                value=st.session_state[f"lvl{i}_cond_count"],
                key=f"lvl{i}_cond_count_ui"
            )
            st.session_state[f"lvl{i}_cond_count"] = ccount

            ops = ["=", "<>", ">", ">=", "<", "<=", "IN", "NOT IN", "CONTAINS", "STARTS WITH", "ENDS WITH", "LIKE"]
            conditions: List[LevelCondition] = []
            for k in range(int(ccount)):
                c1, c2, c3 = st.columns([2, 1, 3])
                with c1:
                    cond_col = st.selectbox(f"Condition {k + 1} column", options=cols, key=f"lvl{i}_cond_{k}_col")
                with c2:
                    cond_op = st.selectbox("Op", options=ops, key=f"lvl{i}_cond_{k}_op")
                with c3:
                    cond_val = st.text_input("Value(s)", placeholder="e.g. MOJ or 2025 or MOJ,NPC", key=f"lvl{i}_cond_{k}_val")
                conditions.append(LevelCondition(cond_col, cond_op, cond_val))

            levels.append(LevelSpec(
                level_no=i,
                object_view=obj_view,
                level_alias_ar=lvl_alias_ar,
                columns=selected_cols,
                col_alias_ar=col_alias_ar,
                conditions=conditions
            ))

        if i < int(st.session_state.level_count):
            with st.expander(f"Relationship between Level {i} -> Level {i + 1}", expanded=(i == 1)):
                rel_views = st.multiselect(
                    "Relationship/View(s) (RelationshipType.[...])",
                    options=st.session_state.rel_views,
                    key=f"hop{i}_rels"
                )

                direction = st.selectbox(
                    "Join direction",
                    options=["Bi-Directional", "From->To", "To->From"],
                    index=0,
                    help=(
                        "Bi-Directional treats the relationship as undirected (both From->To and To->From). "
                        "Use From->To / To->From if you want strict direction."
                    ),
                    key=f"hop{i}_dir"
                )

                hops.append(HopSpec(from_level_no=i, relationship_views=rel_views, join_direction=direction))

    # Validation
    validation_errors = []
    for hop in hops:
        if not hop.relationship_views:
            validation_errors.append(
                f"Please select at least one Relationship between Level {hop.from_level_no} and Level {hop.from_level_no + 1}."
            )

    st.divider()
    st.subheader("3) Generate SQL")

    if validation_errors:
        for msg in validation_errors:
            st.error(msg)
        st.stop()

    cA, cB = st.columns([1, 1])
    with cA:
        distinct = st.checkbox("Use DISTINCT", value=True)
    with cB:
        include_order_by = st.checkbox("Include ORDER BY (for SSMS output)", value=True)

    sql_full, sql_core, wrapper_order_cols, output_fields = generate_sql_parts(
        levels, hops,
        distinct=distinct,
        include_order_by=include_order_by
    )

    st.code(sql_full, language="sql")
    st.download_button(
        "Download SQL",
        data=(sql_full + "\n").encode("utf-8"),
        file_name="iserver_hierarchy.sql",
        mime="text/plain"
    )

    st.divider()
    st.subheader("4) Execute SQL + Export (Excel / RDL)")

    c1, c2, c3 = st.columns([1, 1, 2])
    with c1:
        run_btn = st.button("Execute SQL", type="primary")
    with c2:
        max_rows = st.number_input("Max rows", min_value=10, max_value=200000, value=1000, step=10)
    with c3:
        apply_order_in_preview = st.checkbox(
            "Apply ORDER BY in preview",
            value=True,
            help="ORDER BY is applied outside the wrapper using output column aliases.",
        )

    if run_btn and conn is not None:
        try:
            outer_order_sql = ""
            if apply_order_in_preview and wrapper_order_cols:
                outer_order_sql = "ORDER BY " + ", ".join(wrapper_order_cols)

            safe_sql = f"""
SELECT TOP ({int(max_rows)}) *
FROM (
{sql_core}
) X
{outer_order_sql}
""".strip()

            df = fetch_df(conn, safe_sql)
            st.session_state.last_result_df = df
            st.session_state.last_exec_sql = safe_sql

            st.success(f"Executed ✅ Returned {len(df):,} rows (max {int(max_rows):,}).")

        except Exception as e:
            st.session_state.last_result_df = None
            st.session_state.last_exec_sql = None
            st.error(f"Execution failed: {e}")

    df_last = st.session_state.last_result_df

    if isinstance(df_last, pd.DataFrame) and not df_last.empty:
        st.markdown("### Preview options")

        default_merge_cols = []
        for oc in wrapper_order_cols:
            m = re.match(r"X\.\[(.*)\]$", oc.strip())
            if m:
                default_merge_cols.append(m.group(1))

        merge_cols = st.multiselect(
            "Hierarchy columns to merge (top-down order)",
            options=list(df_last.columns),
            default=[c for c in default_merge_cols if c in df_last.columns]
        )

        preview_rtl = st.checkbox("Preview RTL (Arabic)", value=True)

        merge_preview = st.checkbox("Show merged preview (rowspan HTML)", value=False)
        if merge_preview and merge_cols:
            html = df_to_rowspan_html(df_last, merge_cols, rtl=preview_rtl)
            st.markdown(html, unsafe_allow_html=True)
        else:
            st.dataframe(df_last, use_container_width=True)

        st.markdown("### Excel export")
        merge_excel = st.checkbox(
            "Merge repeated hierarchy cells in Excel",
            value=True,
            help="Merges vertically per selected hierarchy columns (hierarchy-aware)."
        )

        excel_bytes = df_to_excel_bytes(
            df_last,
            sheet_name="Result",
            merge_hierarchy=merge_excel,
            merge_cols=merge_cols
        )

        st.download_button(
            "Export last result to Excel",
            data=excel_bytes,
            file_name="iserver_result.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.markdown("### RDL export (Report Builder)")
        r1, r2, r3, r4 = st.columns([1.2, 1.2, 1.4, 1.2])
        with r1:
            rdl_rtl = st.checkbox("RTL (default)", value=True)
        with r2:
            rdl_repeat_headers = st.checkbox("Repeat table headers", value=True)
        with r3:
            rdl_autofit = st.checkbox("Auto-fit to page width", value=True)
        with r4:
            rdl_sample = st.number_input("Sizing sample rows", min_value=50, max_value=5000, value=2000, step=50)

        server_for_rdl = st.session_state.get("server", "").strip()
        db_for_rdl = st.session_state.get("database", "").strip()
        if override_full and (not server_for_rdl or not db_for_rdl):
            kv = parse_conn_kv(conn_str)
            server_for_rdl = server_for_rdl or kv.get("server", "")
            db_for_rdl = db_for_rdl or kv.get("database", "")

        report_name = st.text_input("Report title", value="iServer Hierarchy Report")

        rdl_bytes = generate_rdl_bytes(
            report_name=report_name,
            sql_query=sql_full,
            fields=list(df_last.columns),
            server=server_for_rdl or "SERVER",
            database=db_for_rdl or "DATABASE",
            group_cols=merge_cols,
            rtl=rdl_rtl,
            df_for_sizing=df_last,
            sample_rows_for_sizing=int(rdl_sample),
            auto_fit_page_width=rdl_autofit,
            repeat_headers=rdl_repeat_headers
        )

        st.download_button(
            "Download RDL (Report Builder Template)",
            data=rdl_bytes,
            file_name="iserver_hierarchy_report.rdl",
            mime="application/xml"
        )

        with st.expander("Show executed SQL (wrapper)", expanded=False):
            st.code(st.session_state.last_exec_sql or "", language="sql")

    elif isinstance(df_last, pd.DataFrame) and df_last.empty:
        st.info("Last execution returned 0 rows — nothing to export yet.")
    else:
        st.info("Run Execute SQL to preview/export results (Excel/RDL).")

else:
    st.info("Connect to the DB first to load ShapeType / RelationshipType views and columns.")
